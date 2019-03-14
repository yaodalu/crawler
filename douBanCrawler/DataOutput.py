#coding:utf-8
from openpyxl import Workbook

#将爬取到的数据写成xlsx格式
class DataOutput(object):
    def output(self,book_lists,book_tag_lists,topath):
        wb=Workbook()
        ws=[]
        save_path= topath
        
        for i in range(len(book_tag_lists)):
            ws.append(wb.create_sheet(title=book_tag_lists[i].decode('utf-8'))) #utf8->unicode 添加sheet
        for i in range(len(book_tag_lists)): 
            ws[i].append([u'序号',u'书名',u'评分',u'评价人数',u'作者',u'出版社'])#每个sheet添加第一行表头
            count=1
            for bl in book_lists[i]:
                ws[i].append([count,bl[0],float(bl[1]),int(bl[2]),bl[3],bl[4]])
                count+=1
            save_path+=('-'+book_tag_lists[i].decode('utf-8'))
        save_path+='.xlsx'
        wb.save(save_path) 
   

    
