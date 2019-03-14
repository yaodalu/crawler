# -*- coding: UTF-8 -*-
import openpyxl
import time
#查看库的对象、属性、函数在cmd窗口Python27目录下键入python -m pydoc openpyxl

ls=[['马坡','接入交换','192.168.1.1','G0/3','AAAA-AAAA-AAAA'],
      ['马坡','接入交换','192.168.1.2','G0/8','BBBB-BBBB-BBBB'],
      ['马坡','接入交换','192.168.1.2','G0/8','CCCC-CCCC-CCCC'],
      ['马坡','接入交换','192.168.1.2','G0/8','DDDD-DDDD-DDDD']]

#不懂
time_format = '%Y-%m-%d__%H:%M:%S'
time_current = time.strftime(time_format)
'''
2019-01-08__10:52:52
'''

def saveToExcel(data,sheetname,wbname):
    print("写入excel: ")
    wb = openpyxl.load_workbook(filename=wbname)
    sheet = wb.active
    print(sheet.max_row)
    row = sheet.max_row + 1
    data_len = row + len(data)

    for data_row in range(row,data_len):
        for data_col1 in range(2,7):
            sheet.cell(row = data_row,column = 1, value = str(time_current))
            sheet.cell(row = data_row,column = data_col1, value =str(data[data_row-data_len][data_col1-2]))

    wb.save(filename = wbname)
    print('保存成功')

saveToExcel(ls,'sheet1',r'C:\Users\1\Desktop\python code\Crawler\1.xlsx')
                            
