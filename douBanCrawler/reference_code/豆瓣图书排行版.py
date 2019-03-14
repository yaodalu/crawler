#-*- coding: UTF-8 -*-

import sys
import time
import urllib
import urllib2
import requests
import numpy as np
from bs4 import BeautifulSoup
from openpyxl import Workbook
import ssl

#Some User Agents
hds=[{'User-Agent':'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'},\
{'User-Agent':'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11'},\
{'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'}]


def book_spider(book_tag):
    page_num=0
    book_list=[]
    try_times=0
    
    while page_num <= 2:
        url='https://www.douban.com/tag/'+urllib.quote(book_tag)+'/book?start='+str(page_num*15) #wd编码 凯哥学堂
        time.sleep(np.random.rand()*5)
    
        try:
            req = urllib2.Request(url, headers=hds[page_num%len(hds)])
            context = ssl.create_default_context()
            source_code = urllib2.urlopen(req,context = context).read()#ssl
            plain_text=str(source_code)   
        except (urllib2.HTTPError, urllib2.URLError), e:
            print e
            continue
        
        soup = BeautifulSoup(plain_text,'lxml')
        list_soup = soup.find('div', {'class': 'mod book-list'})#书目
        
        try_times+=1
        if list_soup==None and try_times<50:
            continue
        elif list_soup==None or len(list_soup)<=1:
            break # Break when no informatoin got after 200 times requesting
        
        for book_info in list_soup.findAll('dd'):
            title = book_info.find('a', {'class':'title'}).string.strip()#strip()移除空格
            desc = book_info.find('div', {'class':'desc'}).string.strip()
            desc_list = desc.split('/') #作者
            book_url = book_info.find('a', {'class':'title'}).get('href')#"https://book.douban.com/subject/26879778/?from=tag_all"
            
            try:
                author_info = u'作者/译者： ' + u'/'.join(desc_list[0:-3])
            except:
                author_info =u'作者/译者： 暂无'
            try:
                pub_info = u'出版信息： ' + u'/'.join(desc_list[-3:])
            except:
                pub_info = u'出版信息： 暂无'
            try:
                rating = book_info.find('span', {'class':'rating_nums'}).string.strip()#9.2
            except:
                rating=u'0.0'
            try:
                #people_num = book_info.findAll('span')[2].string.strip()
                people_num = get_people_num(book_url) 
                people_num = people_num.strip(u'人评价')
            except:
                people_num =u'0'
            
            book_list.append([title,rating,people_num,author_info,pub_info])
            try_times=0 #set 0 when got valid information
        page_num+=1
        print 'Downloading Information From Page %d' % page_num
    return book_list


def get_people_num(url):
    #url='http://book.douban.com/subject/6082808/?from=tag_all' # For Test
    try:
        req = urllib2.Request(url, headers=hds[np.random.randint(0,len(hds))])
        source_code = urllib2.urlopen(req).read()
        plain_text=str(source_code)   
    except (urllib2.HTTPError, urllib2.URLError), e:
        print e
    soup = BeautifulSoup(plain_text,'lxml')
    people_num=soup.find('div',{'class':'rating_sum'}).findAll('span')[1].string.strip()
    return people_num #评分人数


def do_spider(book_tag_lists):
    book_lists=[]
    for book_tag in book_tag_lists:
        book_list=book_spider(book_tag)
        book_list=sorted(book_list,key=lambda x:x[1],reverse=True)#rating 逆排序 从大到小
        book_lists.append(book_list)#[[[序号，书名，评分，评价人数，作者，出版社]]]三层嵌套，最外层是tag,最里层是每本书
    return book_lists


def print_book_lists_excel(book_lists,book_tag_lists):
    wb=Workbook()
    ws=[]
    for i in range(len(book_tag_lists)):
        ws.append(wb.create_sheet(title=book_tag_lists[i].decode('utf-8'))) #utf8->unicode 添加sheet
    for i in range(len(book_tag_lists)): 
        ws[i].append([u'序号',u'书名',u'评分',u'评价人数',u'作者',u'出版社'])#每个sheet添加第一行表头
        count=1
        for bl in book_lists[i]:
            ws[i].append([count,bl[0],float(bl[1]),int(bl[2]),bl[3],bl[4]])
            count+=1
    save_path=r'C:\Users\1\Desktop\python code\Crawler\file\book_list'
    for i in range(len(book_tag_lists)):
        save_path+=('-'+book_tag_lists[i].decode('utf-8'))
    save_path+='.xlsx'
    wb.save(save_path)




if __name__=='__main__':
    #book_tag_lists = ['心理','判断与决策','算法','数据结构','经济','历史']
    book_tag_lists = ['算法','数据结构'']
    #book_tag_lists = ['传记','哲学','编程','创业','理财','社会学','佛教']
    #book_tag_lists = ['思想','科技','科学','web','股票','爱情','两性']
    #book_tag_lists = ['计算机','机器学习','linux','android','数据库','互联网']
    #book_tag_lists = ['数学']
    #book_tag_lists = ['摄影','设计','音乐','旅行','教育','成长','情感','育儿','健康','养生']
    #book_tag_lists = ['商业','理财','管理']  
    #book_tag_lists = ['名著']
    #book_tag_lists = ['科普','经典','生活','心灵','文学']
    #book_tag_lists = ['科幻','思维','金融']
    #book_tag_lists = ['个人管理','时间管理','投资','文化','宗教']
    book_lists=do_spider(book_tag_lists)
    print_book_lists_excel(book_lists,book_tag_lists)
    
